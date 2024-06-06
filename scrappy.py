from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from lxml import html
import requests
import openpyxl

nome = []
descricao = []
preco = []

next_page = '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a'
url = 'https://telefonesimportados.netlify.app/'
cont = 2

servico = Service(executable_path="C:\\Users\\tiago\\anaconda3\\chromedriver.exe")
driver = webdriver.Chrome(service=servico)

while next_page:
    req = requests.get(url)
    driver.get(url)
    url = url + 'shop' + str(cont)
    cont +=1

    tree = html.fromstring(req.text)


    



    encontrei = False
    elem = 1
    
    while not encontrei:
        objeto = tree.xpath(f'/html/body/div[5]/div[2]/div[1]/div[{elem}]/div/div[3]')
        class_objeto = driver.find_element(By.XPATH, f'/html/body/div[5]/div[2]/div[1]/div[{elem}]/div/div[3]' ).get_attribute('class')
        if not 'product-option-shop' in class_objeto:
            encontrei = True
        else:
            nome[elem - 1] = nome.append(tree.xpath(f'/html/body/div[5]/div[2]/div[1]/div[{elem}]/div/h2/a/text()')[0])
            # descricao[i] = descricao.append(tree.xpath('//a[@class="shelf-default__product-name"]/text()'))
            preco[elem - 1] = preco.append(tree.xpath(f'/html/body/div[5]/div[2]/div[1]/div[{elem}]/div/div[2]/ins/text()')[0])
        elem += 1









index=1
planilha = openpyxl.Workbook()
celulares = planilha['Sheet']
celulares.title = 'Celulares'
celulares['A1'] = 'Nome'
celulares['B1'] = 'Preço'

for a,b in zip(nome,preco):
    celulares.cell(column=1,row=index,value=a)
    celulares.cell(column=2,row=index,value=b)
    index +=1

planilha.save('planilha_cel.xlsx')

print(nome)
print(preco)
print()


